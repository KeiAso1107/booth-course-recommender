"""
Data processing pipeline for Chicago Booth Course Recommender.
Loads bid history and course evaluation data, cleans, and outputs unified DataFrames.
"""

import pandas as pd
import numpy as np
import openpyxl
import re
import pickle
from pathlib import Path

DATA_DIR = Path(__file__).parent

# Quarter ordering for sorting
QUARTER_ORDER = {"Winter": 0, "Spring": 1, "Summer": 2, "Autumn": 3}


def parse_price(val):
    """Parse a price value, handling CLO, N/A, None, etc."""
    if val is None or val == "N/A" or val == "":
        return np.nan
    if isinstance(val, str):
        val = val.strip().replace(",", "").replace("$", "")
        if val.upper() in ("CLO", "CLOSED", "N/A", ""):
            return np.nan
        try:
            return float(val)
        except ValueError:
            return np.nan
    return float(val)


def parse_seats(val):
    """Parse seats available, handling CLO (closed = 0)."""
    if val is None or val == "N/A" or val == "":
        return np.nan
    if isinstance(val, str):
        val = val.strip()
        if val.upper() in ("CLO", "CLOSED"):
            return 0
        try:
            return int(float(val))
        except ValueError:
            return np.nan
    return int(val)


def parse_enrollment(val):
    """Parse enrollment count."""
    if val is None or val == "N/A" or val == "":
        return np.nan
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return np.nan


def load_bid_data():
    """
    Load all sheets from Course Price History.xlsx into a unified DataFrame.

    Returns a DataFrame with columns:
        course_code, title, quarter, year, day_time, instructor,
        phase, enrollment, seats_available, clearing_price
    Each row = one course-section × one phase.
    """
    filepath = DATA_DIR / "Course Price History.xlsx"
    wb = openpyxl.load_workbook(filepath, read_only=True)

    all_records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else "" for h in rows[0]]

        # Identify phase groups: triplets of (Enrollment, Seats, Price)
        # Pattern: columns come in groups of 3 after the first 6 base columns
        base_cols = 6  # Course, Title, Quarter, Year, Day and Time, Instructor

        # Extract phase info from headers
        phases = []
        i = base_cols
        while i + 2 < len(headers):
            enroll_h = headers[i]
            seats_h = headers[i + 1]
            price_h = headers[i + 2]

            # Extract phase name from the price column header
            phase_name = price_h.replace(" Price", "").strip()
            # Normalize phase names
            phase_name = (phase_name
                         .replace("Phase 25", "Phase 2.5")  # Winter 2025 anomaly
                         .replace("Phase 1 New Students", "New Students P1")
                         .replace("Phase 2 New Students", "New Students P2")
                         .replace("New Students", "New Students"))

            phases.append((phase_name, i, i + 1, i + 2))
            i += 3

        # Process data rows
        for row in rows[1:]:
            if row[0] is None:
                continue

            course_code = str(row[0]).strip()
            title = str(row[1]).strip() if row[1] else ""
            quarter = str(row[2]).strip() if row[2] else ""
            year = int(row[3]) if row[3] else 0
            day_time = str(row[4]).strip() if row[4] else ""
            instructor = str(row[5]).strip() if row[5] else ""

            for phase_name, ei, si, pi in phases:
                enrollment = parse_enrollment(row[ei] if ei < len(row) else None)
                seats = parse_seats(row[si] if si < len(row) else None)
                price = parse_price(row[pi] if pi < len(row) else None)

                all_records.append({
                    "course_code": course_code,
                    "title": title,
                    "quarter": quarter,
                    "year": year,
                    "day_time": day_time,
                    "instructor": instructor,
                    "phase": phase_name,
                    "enrollment": enrollment,
                    "seats_available": seats,
                    "clearing_price": price,
                })

    wb.close()

    df = pd.DataFrame(all_records)

    # Parse course department and number
    df["dept_code"] = df["course_code"].apply(lambda x: re.match(r"(\d{5})", x).group(1) if re.match(r"(\d{5})", x) else x)
    df["section"] = df["course_code"].apply(lambda x: x.split("-")[-1] if "-" in x else "")

    # Time slot features
    df["is_evening"] = df["day_time"].str.contains("6:00 pm|7:00 pm", case=False, na=False)
    df["is_weekend"] = df["day_time"].str.contains("Saturday|Sunday", case=False, na=False)

    # Day of week
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    for day in days:
        df[f"day_{day.lower()}"] = df["day_time"].str.contains(day, case=False, na=False).astype(int)

    # Quarter numeric for sorting
    df["quarter_num"] = df["quarter"].map(QUARTER_ORDER)
    df["term_order"] = df["year"] * 10 + df["quarter_num"]

    # Is section closed (seats = 0)?
    df["is_closed"] = (df["seats_available"] == 0).astype(int)

    # Capacity estimate (enrollment + seats at Phase 1)
    phase1 = df[df["phase"] == "Phase 1"].copy()
    phase1["capacity"] = phase1["enrollment"] + phase1["seats_available"].fillna(0)
    capacity_map = phase1.groupby(["course_code", "quarter", "year"])["capacity"].first()
    df = df.merge(
        capacity_map.reset_index().rename(columns={"capacity": "est_capacity"}),
        on=["course_code", "quarter", "year"],
        how="left"
    )

    # Fill ratio (enrollment / capacity)
    df["fill_ratio"] = np.where(
        df["est_capacity"] > 0,
        df["enrollment"] / df["est_capacity"],
        np.nan
    )

    return df


def load_evaluation_data():
    """
    Load course evaluation data.

    Returns a DataFrame with columns:
        course_code, title, instructor_first, instructor_last, term,
        invited, respondents, response_rate, hours_per_week,
        clarity, interesting, useful_tools, how_much_got, recommend
    """
    filepath = DATA_DIR / "Booth_MBA_Course_Evaluation_Data.xlsx"
    df = pd.read_excel(filepath, sheet_name="Sheet1")

    # Rename columns (handle varying column count)
    col_names = [
        "course_code", "title", "instructor_first", "instructor_last", "term",
        "invited", "respondents", "response_rate", "hours_per_week",
        "clarity", "interesting", "useful_tools", "how_much_got", "recommend",
    ]
    # Trim or pad to match actual columns
    if len(df.columns) > len(col_names):
        df = df.iloc[:, :len(col_names)]
    df.columns = col_names[:len(df.columns)]

    # Clean course code: "33501 01" -> "33501-01"
    df["course_code_clean"] = df["course_code"].astype(str).str.strip().str.replace(r"\s+", "-", regex=True)
    df["dept_code"] = df["course_code_clean"].apply(lambda x: x.split("-")[0])

    # Parse term -> quarter + year
    def parse_term(t):
        parts = str(t).strip().split()
        if len(parts) == 2:
            return parts[0], int(parts[1])
        return t, 0

    term_parsed = df["term"].apply(parse_term)
    df["quarter"] = term_parsed.apply(lambda x: x[0])
    df["year"] = term_parsed.apply(lambda x: x[1])

    # Overall score (average of the 5 rating dimensions)
    rating_cols = ["clarity", "interesting", "useful_tools", "how_much_got", "recommend"]
    df["overall_score"] = df[rating_cols].mean(axis=1)

    # Instructor full name
    df["instructor"] = df["instructor_last"].astype(str) + ", " + df["instructor_first"].astype(str)

    return df


def build_course_features():
    """
    Build a unified course feature table combining bid history and evaluations.
    Returns a course-level summary DataFrame.
    """
    bid_df = load_bid_data()
    eval_df = load_evaluation_data()

    # --- Bid summary per course (across all quarters) ---
    # Focus on Phase 1 and Phase 2 prices (most relevant for bidding strategy)
    bid_p1 = bid_df[bid_df["phase"] == "Phase 1"].copy()
    bid_p2 = bid_df[bid_df["phase"] == "Phase 2"].copy()

    # Course-level bid stats (using dept_code to match across sections)
    def bid_stats(phase_df, prefix):
        stats = phase_df.groupby("dept_code").agg(
            **{
                f"{prefix}_price_mean": ("clearing_price", "mean"),
                f"{prefix}_price_max": ("clearing_price", "max"),
                f"{prefix}_price_median": ("clearing_price", "median"),
                f"{prefix}_price_std": ("clearing_price", "std"),
                f"{prefix}_fill_mean": ("fill_ratio", "mean"),
                f"{prefix}_closed_rate": ("is_closed", "mean"),
                f"{prefix}_count": ("clearing_price", "count"),
            }
        ).reset_index()
        return stats

    p1_stats = bid_stats(bid_p1, "p1")
    p2_stats = bid_stats(bid_p2, "p2")

    # Course metadata (latest offering)
    latest = bid_df.sort_values("term_order", ascending=False).groupby("dept_code").first().reset_index()
    course_meta = latest[["dept_code", "title", "est_capacity"]].copy()

    # --- Evaluation summary per course ---
    eval_summary = eval_df.groupby("dept_code").agg(
        eval_clarity=("clarity", "mean"),
        eval_interesting=("interesting", "mean"),
        eval_useful=("useful_tools", "mean"),
        eval_got_out=("how_much_got", "mean"),
        eval_recommend=("recommend", "mean"),
        eval_overall=("overall_score", "mean"),
        eval_hours=("hours_per_week", "mean"),
        eval_count=("course_code_clean", "count"),
    ).reset_index()

    # --- Evaluation titles (for courses only in eval data) ---
    eval_titles = eval_df.groupby("dept_code")["title"].first().reset_index()
    eval_titles = eval_titles.rename(columns={"title": "eval_title"})

    # --- Merge ---
    features = course_meta.merge(p1_stats, on="dept_code", how="outer")
    features = features.merge(p2_stats, on="dept_code", how="outer")
    features = features.merge(eval_summary, on="dept_code", how="outer")
    features = features.merge(eval_titles, on="dept_code", how="left")

    # Fill missing titles from eval data
    features["title"] = features["title"].fillna(features["eval_title"])
    features = features.drop(columns=["eval_title"])

    return features, bid_df, eval_df


def load_course_list():
    """
    Load 2025-2026 course list.
    Returns DataFrame with columns:
        quarter, title, course_code, section, dept_code, program,
        faculty, schedule, capacity_str, building, location
    """
    filepath = DATA_DIR / "Course List.xlsx"
    df = pd.read_excel(filepath)
    df.columns = [
        "quarter", "title", "course_code", "program",
        "faculty", "schedule", "capacity_str", "building", "location",
    ]

    # Parse course code
    df["dept_code"] = df["course_code"].str.split("-").str[0].str.strip()
    df["section"] = df["course_code"].str.split("-").str[1].str.strip()

    # Normalize faculty name to "Last, First" for matching with bid data
    def normalize_faculty(name):
        if pd.isna(name) or name.strip() == "":
            return ""
        parts = name.strip().split()
        if len(parts) >= 2:
            return parts[-1] + ", " + " ".join(parts[:-1])
        return name
    df["instructor"] = df["faculty"].apply(normalize_faculty)

    return df


def get_instructor_stats(eval_df):
    """Get instructor-level aggregated ratings."""
    return eval_df.groupby(["instructor_last", "instructor_first"]).agg(
        courses_taught=("course_code_clean", "nunique"),
        sections_taught=("course_code_clean", "count"),
        avg_clarity=("clarity", "mean"),
        avg_interesting=("interesting", "mean"),
        avg_useful=("useful_tools", "mean"),
        avg_got_out=("how_much_got", "mean"),
        avg_recommend=("recommend", "mean"),
        avg_overall=("overall_score", "mean"),
        avg_hours=("hours_per_week", "mean"),
    ).reset_index().sort_values("avg_overall", ascending=False)


def get_price_trends(bid_df, dept_code):
    """Get price trend over time for a specific course."""
    course = bid_df[bid_df["dept_code"] == dept_code].copy()
    course = course.sort_values("term_order")

    # Pivot by phase
    p2 = course[course["phase"] == "Phase 2"].groupby(["quarter", "year", "term_order"]).agg(
        avg_price=("clearing_price", "mean"),
        max_price=("clearing_price", "max"),
        avg_fill=("fill_ratio", "mean"),
    ).reset_index().sort_values("term_order")

    return p2


if __name__ == "__main__":
    print("Loading and processing data...")
    features, bid_df, eval_df = build_course_features()

    print(f"\nBid data: {len(bid_df)} rows")
    print(f"Evaluation data: {len(eval_df)} rows")
    print(f"Unique courses (bid): {bid_df['dept_code'].nunique()}")
    print(f"Unique courses (eval): {eval_df['dept_code'].nunique()}")
    print(f"Course features table: {len(features)} courses")

    print("\n=== Top 20 most expensive courses (Phase 2 avg) ===")
    top = features.nlargest(20, "p2_price_mean")[["dept_code", "title", "p2_price_mean", "p2_price_max", "p2_closed_rate", "eval_overall", "eval_hours"]]
    print(top.to_string(index=False))

    print("\n=== Top 20 highest rated courses ===")
    top_rated = features.dropna(subset=["eval_overall"]).nlargest(20, "eval_overall")[["dept_code", "title", "eval_overall", "eval_recommend", "eval_hours", "p2_price_mean"]]
    print(top_rated.to_string(index=False))

    # Save processed data
    bid_df.to_csv(DATA_DIR / "bid_data_processed.csv", index=False)
    eval_df.to_csv(DATA_DIR / "eval_data_processed.csv", index=False)
    features.to_csv(DATA_DIR / "course_features.csv", index=False)
    print("\nProcessed data saved to CSV files.")
